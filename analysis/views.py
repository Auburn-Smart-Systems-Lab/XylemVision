from django.shortcuts import render
from django.http import HttpResponse, HttpResponseBadRequest
from PIL import Image
from io import BytesIO
import base64
import openpyxl
from openpyxl.styles import PatternFill
from .engine import progressive_yolo_sam

_last_analysis_cache = {}

def pil_to_base64(img):
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    encoded = base64.b64encode(buffer.getvalue()).decode()
    return f"data:image/png;base64,{encoded}"

def root_analysis_view(request):
    context = {}
    results = []

    if request.method == 'POST' and request.FILES.getlist('image'):
        images = request.FILES.getlist('image')
        results.clear()

        for image_file in images:
            image = Image.open(image_file).convert('RGB')
            result, original_img, overlay_img = progressive_yolo_sam(image)

            xylem_details = result['metrics'].get('xylem_details', [])
            colours = [c for c in result.get('colours', []) if c.get('class') == 'Xylem']

            merged = []
            extra_metrics = []
            extra_colours = []

            max_len = max(len(xylem_details), len(colours))

            for i in range(max_len):
                detail = xylem_details[i] if i < len(xylem_details) else None
                color = colours[i] if i < len(colours) and colours[i]['inst'] == i else None

                if detail and color:
                    merged.append({
                        'instance': i,
                        **detail,
                        'rgb': color['rgb']
                    })
                elif detail:
                    extra_metrics.append({'instance': i, **detail})
                elif color:
                    extra_colours.append({'instance': i, 'rgb': color['rgb']})

            analysis = {
                'file': image_file.name,
                'original_image': pil_to_base64(original_img),
                'overlay_image': pil_to_base64(overlay_img),
                'n_xylem': result['n_xylem'],
                'n_vb': result['n_vb'],
                'n_root': result['n_root'],
                'metrics': result['metrics'],
                'merged_xylem': merged,
                'extra_metrics': extra_metrics,
                'extra_colours': extra_colours,
            }

            results.append(analysis)
            _last_analysis_cache[image_file.name] = analysis

        context['results'] = results

    return render(request, 'upload.html', context)

# Helper: create XLSX workbook for one analysis result
def generate_xlsx_for_result(result):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Summary Metrics
    ws.append(["Metric", "Value"])
    ws.append(["Xylem Count", result['n_xylem']])
    ws.append(["Vascular Bundle Total Area", result['metrics'].get('vb_total_area', '')])
    ws.append(["Vascular Bundle Max Diameter", result['metrics'].get('vb_max_diameter', '')])
    ws.append(["Root Total Area", result['metrics'].get('root_total_area', '')])
    ws.append(["Root Max Diameter", result['metrics'].get('root_max_diameter', '')])
    ws.append([])

    # Merged Xylem Details with Colors
    merged = result.get('merged_xylem', [])
    if merged:
        ws.append(["Xylem Details"])
        headers = [key for key in merged[0].keys() if key not in ('instance', 'rgb')]
        headers.append("Color (RGB)")
        ws.append(headers)

        for row in merged:
            row_values = [row[key] for key in headers if key != "Color (RGB)"]
            row_values.append(str(tuple(row['rgb'])))
            ws.append(row_values)

            # Apply fill color to the last cell
            fill_color = openpyxl.styles.PatternFill(
                start_color="{:02X}{:02X}{:02X}".format(*row['rgb']),
                end_color="{:02X}{:02X}{:02X}".format(*row['rgb']),
                fill_type="solid"
            )
            cell = ws.cell(row=ws.max_row, column=len(headers))
            cell.fill = fill_color

        ws.append([])

    # Extra unmatched metrics
    extra_metrics = result.get('extra_metrics', [])
    if extra_metrics:
        ws.append(["Unmatched Xylem Metrics"])
        headers = [key for key in extra_metrics[0].keys() if key != 'instance']
        ws.append(headers)
        for row in extra_metrics:
            ws.append([row[key] for key in headers])
        ws.append([])

    # Extra unmatched colors
    extra_colours = result.get('extra_colours', [])
    if extra_colours:
        ws.append(["Unmatched Overlay Colors"])
        ws.append(["Color (RGB)"])
        for color in extra_colours:
            ws.append([str(tuple(color['rgb']))])
            fill_color = openpyxl.styles.PatternFill(
                start_color="{:02X}{:02X}{:02X}".format(*color['rgb']),
                end_color="{:02X}{:02X}{:02X}".format(*color['rgb']),
                fill_type="solid"
            )
            cell = ws.cell(row=ws.max_row, column=1)
            cell.fill = fill_color

    return wb

# Download XLSX for a single analysis
def download_xlsx(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("Invalid request method")

    filename = request.POST.get('filename')
    if not filename or filename not in _last_analysis_cache:
        return HttpResponseBadRequest("No analysis found for that file")

    result = _last_analysis_cache[filename]
    wb = generate_xlsx_for_result(result)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}_analysis.xlsx"'
    wb.save(response)
    return response

# Download XLSX for all analyses combined
def download_all_xlsx(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("Invalid request method")

    filenames = request.POST.getlist('filenames')
    if not filenames:
        return HttpResponseBadRequest("No files selected")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for filename in filenames:
        if filename not in _last_analysis_cache:
            continue
        result = _last_analysis_cache[filename]

        ws = wb.create_sheet(title=filename[:31])  # max sheet title length is 31
        # Summary Metrics
        ws.append(["Metric", "Value"])
        ws.append(["Xylem Count", result['n_xylem']])
        ws.append(["Vascular Bundle Total Area", result['metrics'].get('vb_total_area', '')])
        ws.append(["Vascular Bundle Max Diameter", result['metrics'].get('vb_max_diameter', '')])
        ws.append(["Root Total Area", result['metrics'].get('root_total_area', '')])
        ws.append(["Root Max Diameter", result['metrics'].get('root_max_diameter', '')])
        ws.append([])

        # Merged Xylem Details
        merged = result.get('merged_xylem', [])
        if merged:
            ws.append(["Xylem Details"])
            headers = [key for key in merged[0].keys() if key not in ('instance', 'rgb')]
            headers.append("Color (RGB)")
            ws.append(headers)

            for row in merged:
                row_values = [row[key] for key in headers if key != "Color (RGB)"]
                row_values.append(str(tuple(row['rgb'])))
                ws.append(row_values)

                fill_color = openpyxl.styles.PatternFill(
                    start_color="{:02X}{:02X}{:02X}".format(*row['rgb']),
                    end_color="{:02X}{:02X}{:02X}".format(*row['rgb']),
                    fill_type="solid"
                )
                cell = ws.cell(row=ws.max_row, column=len(headers))
                cell.fill = fill_color

            ws.append([])

        # Extra unmatched metrics
        extra_metrics = result.get('extra_metrics', [])
        if extra_metrics:
            ws.append(["Unmatched Xylem Metrics"])
            headers = [key for key in extra_metrics[0].keys() if key != 'instance']
            ws.append(headers)
            for row in extra_metrics:
                ws.append([row[key] for key in headers])
            ws.append([])

        # Extra unmatched colors
        extra_colours = result.get('extra_colours', [])
        if extra_colours:
            ws.append(["Unmatched Overlay Colors"])
            ws.append(["Color (RGB)"])
            for color in extra_colours:
                ws.append([str(tuple(color['rgb']))])
                fill_color = openpyxl.styles.PatternFill(
                    start_color="{:02X}{:02X}{:02X}".format(*color['rgb']),
                    end_color="{:02X}{:02X}{:02X}".format(*color['rgb']),
                    fill_type="solid"
                )
                cell = ws.cell(row=ws.max_row, column=1)
                cell.fill = fill_color

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="all_analysis_combined.xlsx"'
    wb.save(response)
    return response
